import openpyxl
import os

wb = openpyxl.Workbook()

#プレイヤー名入力
print("player name please")
print("input example(exp exp exp)\n")
name = list(map(str, input().split()))
#例外処理
if len(name) == 0:
  print("ERROR:No input")
  exit()

#各ルール別にスコア表を生成
#COUNT_UP()
#ZERO_ONE()
#CRICKET()

wb.save(os.environ['HOME']+'/Desktop/score_sheet.xlsx')

def COUNT_UP():#count-up用スコア表
#スコア項目作成
    sheet = wb.worksheets[0]
    sheet.title = 'COUNT-UP'
    sheet['A1'] = "PLAYER NAME"
    sheet['J1'] = "SCORE"
    count = 1
    for i in range(66,74):
      col = chr(i)+"1"
      sheet[col] = str(count)+"R"
      count += 1
#名前入力&セル設定
    for i in range(len(name)):
      row = "A"+str(i+2)
      sheet[row] = name[i]
      row = "K"+str(i+2)
      com_area = "B"+str(i+2)+":J"+str(i+2)
      sheet[row] = '=SUM('+com_area+')'

def ZERO_ONE():#zero-one用スコア表
    wb.create_sheet()
    sheet = wb.worksheets[1]
    sheet.title = 'ZERO-ONE'
    sheet['A1'] = "PLAYER NAME"
    sheet['B1'] = "SCORE"
    count = 1
    for i in range(67,90):
      col = chr(i)+"1"
      sheet[col] = str(count)+"R"
      count += 1

    for i in range(len(name)):
      row = "A"+str(i+2)
      sheet[row] = name[i]
      row = "B"+str(i+2)
      com_area = "C"+str(i+2)+":Z"+str(i+2)
      #設定持ち点に応じて変数変更(301,501,701,1001)
      Score = str(501)
      sheet[row] = '='+Score+'-SUM'+'('+com_area+')'

def CRICKET():#Cricket用命中回数メモ
    wb.create_sheet()
    sheet = wb.worksheets[2]
    sheet.title = 'CRICKET'
    sheet['A1'] = "PLAYER NAME ＼ NUMBER"
    sheet['V1'] = 50
    count = 1
    for i in range(66,86):
      col = chr(i)+"1"
      sheet[col] = count
      count += 1

    for i in range(len(name)):
      row = "A"+str(i+2)
      sheet[row] = name[i]

#各ルール別にスコア表を生成
COUNT_UP()
ZERO_ONE()
CRICKET()

wb.save(os.environ['HOME']+'/Desktop/score_sheet.xlsx')

